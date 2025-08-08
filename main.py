import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_bdr_template(filename="BDR_Template.xlsx"):
    # Создаем новый workbook
    wb = Workbook()

  # Создаем лист для БДР
    ws = wb.create_sheet("БДР", 0)

    # Заголовки
    headers = [
        "Категория", "Статья", "Ед. изм.", "План (месяц)", 
        "Факт (месяц)", "Отклонение", "План (год)", "Факт (год)", "Отклонение (год)"
    ]

    # Заполняем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Пример данных для доходов
    income_data = [
        ["Доходы", "Выручка от продаж", "руб.", "", "", "", "", "", ""],
        ["Доходы", "Прочие доходы", "руб.", "", "", "", "", "", ""],
        ["Доходы", "Итого доходы", "руб.", "", "", "", "", "", ""],
    ]

    # Пример данных для расходов
    expense_data = [
        ["Расходы", "Себестоимость продаж", "руб.", "", "", "", "", "", ""],
        ["Расходы", "Коммерческие расходы", "руб.", "", "", "", "", "", ""],
        ["Расходы", "Управленческие расходы", "руб.", "", "", "", "", "", ""],
        ["Расходы", "Амортизация", "руб.", "", "", "", "", "", ""],
        ["Расходы", "Прочие расходы", "руб.", "", "", "", "", "", ""],
        ["Расходы", "Итого расходы", "руб.", "", "", "", "", "", ""],
    ]

    # Заполняем данные
    row_num = 2
    for data_row in income_data:
        for col_num, value in enumerate(data_row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        row_num += 1

    # Пустая строка
    row_num += 1

    # Заполняем расходы
    for data_row in expense_data:
        for col_num, value in enumerate(data_row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        row_num += 1

    # Формулы для расчета отклонений
    for row in range(2, row_num):
        # Отклонение за месяц (руб.)
        ws.cell(row=row, column=6).value = f"=E{row}-D{row}"
        # Отклонение за год (руб.)
        ws.cell(row=row, column=9).value = f"=H{row}-G{row}"

    # Итоговая строка
    ws.cell(row=row_num+1, column=1, value="Прибыль/Убыток")
    ws.cell(row=row_num+1, column=4).value = "=SUMIF(A:A,\"Итого доходы\",D:D)-SUMIF(A:A,\"Итого расходы\",D:D)"
    ws.cell(row=row_num+1, column=5).value = "=SUMIF(A:A,\"Итого доходы\",E:E)-SUMIF(A:A,\"Итого расходы\",E:E)"
    ws.cell(row=row_num+1, column=6).value = f"=E{row_num+1}-D{row_num+1}"
    ws.cell(row=row_num+1, column=7).value = "=SUMIF(A:A,\"Итого доходы\",G:G)-SUMIF(A:A,\"Итого расходы\",G:G)"
    ws.cell(row=row_num+1, column=8).value = "=SUMIF(A:A,\"Итого доходы\",H:H)-SUMIF(A:A,\"Итого расходы\",H:H)"
    ws.cell(row=row_num+1, column=9).value = f"=H{row_num+1}-G{row_num+1}"

    # Стилизация итоговой строки
    for col in range(1, 10):
        cell = ws.cell(row=row_num+1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Автофильтр
    ws.auto_filter.ref = f"A1:I{row_num+1}"

    # Ширина колонок
    column_widths = [15, 25, 10, 15, 15, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Создаем второй лист с пояснениями
    ws_help = wb.create_sheet("Пояснения", 1)

    help_data = [
        ["ПОЯСНЕНИЯ К ШАБЛОНУ БДР", ""],
        ["", ""],
        ["1. Категория", "Группировка статей (Доходы/Расходы)"],
        ["2. Статья", "Конкретная статья доходов или расходов"],
        ["3. Ед. изм.", "Единица измерения (руб., шт., % и т.д.)"],
        ["4. План (месяц)", "Плановые значения на месяц"],
        ["5. Факт (месяц)", "Фактические значения на месяц"],
        ["6. Отклонение", "Факт - План (месяц)"],
        ["7. План (год)", "Плановые значения на год"],
        ["8. Факт (год)", "Фактические значения на год"],
        ["9. Отклонение (год)", "Факт - План (год)"],
        ["", ""],
        ["Формат ввода:", "Числовые значения вводятся без пробелов и символов (например: 1000000)"],
        ["Цветовая схема:", "Синий - заголовки, Серый - итоги, Белый - данные"],
    ]

    for row_num, row_data in enumerate(help_data, 1):
        for col_num, value in enumerate(row_data, 1):
            ws_help.cell(row=row_num, column=col_num, value=value)

    # Стилизация заголовка пояснений
    ws_help.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_help.cell(row=1, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_help.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")

    # Ширина колонок для пояснений
    ws_help.column_dimensions['A'].width = 20
    ws_help.column_dimensions['B'].width = 50

    # Сохраняем файл
    wb.save(filename)
    print(f"Шаблон БДР успешно создан: {filename}")

# Создаем шаблон
if __name__ == "__main__":
    create_bdr_template()

